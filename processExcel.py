from openpyxl import load_workbook

import json

class DishInfo(object):
    def __init__(self):
        self.dish_description = ""
        self.allergy_flags = set()

    def __repr__(self):
        return repr((self.dish_description, self.allergy_flags))

def process(excelFile):
    wb = load_workbook(excelFile,read_only=True,data_only=True)
    ws = wb["Chef- Weekly"]

    def get_allergens(index):
        allergens = ["vegan","vegetarian","gluten","soy","milk","eggs","fish","shellfish","peanuts","treenuts"]

        current_allergens = set()

        for i, allergen in enumerate(allergens):
            value = ws.cell(row = index, column = 13+2*i)
            if value is not None and value.value == True:
                current_allergens.add(allergen)

        return current_allergens    



    def get_dish(index):
        dish_description = ws.cell(row = index, column = 2).value

        if dish_description is None:
            return None


        info = DishInfo()
        info.dish_description = dish_description
        info.allergy_flags = get_allergens(index)
        return info
        

    def get_meal(startingPosition):
        if startingPosition is None:
            return None

        return [get_dish(i+startingPosition+1) for i in range(7) if get_dish(i+startingPosition+1) is not None]

    lunchStarts = [1, 17, 33, 49, 65, 81, 89]

    lunches = map(get_meal,lunchStarts)

    dinnerStart = [9, 25, 41, 57, 73, None, 97]
    dinners = map(get_meal,dinnerStart)



    def toDict(aList):
        return {i:elem for i,elem in enumerate(aList) if elem is not None}

    return {
        "base_date": ws.cell(row=1,column=1).value,
        "lunch": toDict(lunches),
        "dinner": toDict(dinners)
    }



if __name__ == "__main__":
    print process(open("/home/ethan/Downloads/Seibel.xlsm","rb"))['dinner'][1]











